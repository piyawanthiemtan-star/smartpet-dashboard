# -*- coding: utf-8 -*-
"""
ML2 Daily Warehouse/Shipping Report  (v1 - local test)
-------------------------------------------------------
อ่านไฟล์ backup ล่าสุดของ POS Onepoint (ML2) แล้วสร้างรายงาน 2 ส่วน:
  1) วันนี้ขายอะไรออกไปบ้าง   -> output/today_sales.csv
  2) ต้องเบิก/เติมอะไรบ้าง     -> output/requisition.csv

เกณฑ์เบิก (ตกลงกับคุณ Piyawan):
  - คิดจากยอดขาย "ของวันนี้วันเดียว"
  - ขายกี่ชิ้น เบิกเท่านั้น (ขั้นต่ำ 1)
  - ยึด single_barcode เป็นตัวจริง แล้วเอา "ชื่อ + ตัวคูณ" จาก Master_Multiplier.xlsx
    (เพราะชื่อใน POS สาขาบางตัวไม่ตรงกับ Master)
  - แนบคะแนน ABC + ธง "ขายดีแต่ของเหลือน้อย" ไว้จัดลำดับความสำคัญ

รันด้วย:  python ml2_report.py
"""
import csv
import glob
import html
import json
import math
import os
import re
import sqlite3
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")

HERE = os.path.dirname(os.path.abspath(__file__))
# env override ทุก path — ให้รันบน GitHub Actions ได้ (เครื่อง Owner ไม่ตั้ง env = พฤติกรรมเดิมทุกอย่าง)
BACKUP_DIR = os.environ.get("ML2R_BACKUP_DIR", HERE)             # ที่เก็บ backup ML2 (Backup-*.db)
OUTPUT_DIR = os.environ.get("ML2R_OUT", os.path.join(HERE, "output"))
MASTER_CSV = os.environ.get("ML2R_MASTER_CSV", os.path.join(HERE, "master_pack_mapping.csv"))   # ไฟล์แก้มือ (override มาสเตอร์)
MASTER_XLSX = os.environ.get("ML2R_MASTER_XLSX", r"D:\1. SmartPet AI Framework\SmartPetData\Master\OnePoint\Master_Multiplier.xlsx")
ML3_DIR = os.environ.get("ML2R_ML3_DIR", r"D:\1. SmartPet AI Framework\SmartPetBackup\Daily")   # ที่เก็บ backup ML3 (คลังต้นทาง)
SEQ_FILE = os.path.join(HERE, ".doc_seq.json")  # เก็บหมายเลขรันล่าสุดของแต่ละวัน

# ชื่อหน่วยที่ใช้เรียกแพ็คตามจำนวนต่อแพ็ค
PACK_UNIT_BY_MULT = {12: "โหล", 24: "ลัง"}

# กฎเบิกสินค้ายกโหล x12 [Piyawan 2026-07-30] (ใช้เฉพาะ x12 ก่อน)
DOZEN_FAST_SALE = 7   # ขาย >= 7/วัน = ขายดี -> เบิก ceil(ขาย/12) โหล
DOZEN_LOW_STOCK = 4   # ขาย < 7 แต่สต๊อกเหลือ < 4 -> เบิก 1 โหล ; ถ้าของยังพอ = ไม่เบิก

# กฎเบิกสินค้าแพ็ค x20 [Piyawan ยืนยัน 2026-08-04] — แพ็คใหญ่เกือบเท่าตัวของ x12
# วิเคราะห์จากข้อมูลจริง (71 SKU): ส่วนใหญ่ขายช้ามาก (<0.5 ชิ้น/วันเฉลี่ย) ปัดเป็นแพ็คเต็มทุกครั้งจะกองของเกิน
X20_FAST_SALE = 8    # ขาย >= 8 ชิ้น/วัน = ขายดีผิดปกติ -> เบิก ceil(ขาย/20) แพ็ค
X20_LOW_STOCK = 3    # ขาย < 8 แต่สต๊อกเหลือ <= 3 -> เบิก 1 แพ็ค (กันของหมด)
                      # ขาย < 8 และสต๊อกยังพอ (>3) -> เบิกเป็น "ชิ้น" เท่าที่ขายจริง ไม่ปัดเป็นแพ็ค (กันของกองเกิน)

# กฎ "ของยังพอ ไม่ต้องเบิก" [Owner เคาะ 3 วัน 2026-08-10 + มติประชุมทีมคลัง 10 ส.ค.]
# แก้ต้นเหตุสต๊อก C บวม (เคยถึง 66%): ML3 วิ่งส่งของให้ ML2 ทุกเช้า (lead time 1 วัน)
# ML2 จึงไม่ต้องถือสต๊อกกันชนเยอะ — คงเหลือพอขายเกินเกณฑ์ -> ไม่เบิก ปล่อยขายกองเดิมก่อน
# (รายการที่ถูกข้ามเก็บใน output/skipped_enough_stock.csv)
#
# วิธีคิด "พอขายกี่วัน" (มติทีมคลัง: ใช้ฤดูกาล + ABC ประกอบ):
# - ฤดูกาล: ใช้ยอดขายเฉลี่ย/วัน ตัวที่แรงกว่า ระหว่างเฉลี่ย 30 วัน กับเฉลี่ย 7 วันล่าสุด
#   (ช่วงเข้าหน้าขาย ยอด 7 วันจะพุ่งก่อน -> ระบบเห็นแล้วเบิกไวขึ้นเอง)
# - ABC: คลาส A/B ใช้เกณฑ์ 3 วันตามที่ Owner เคาะ · คลาส C ใช้ 2 วัน (ตัดแรงกว่า
#   เพราะ C คือกลุ่มที่บวมอยู่ ต้องระบายก่อน) — ปรับตัวเลขได้ที่ dict ข้างล่าง
REQ_MAX_COVER_DAYS = 3
REQ_COVER_DAYS_BY_ABC = {"A": 3, "B": 3, "C": 2, "-": 3}
COVER_WINDOW_DAYS = 30
COVER_RECENT_DAYS = 7

# ใบเบิกย้อนหลังกี่วัน (ดรอปดาวน์เลือกวันที่บนเว็บ /requisition) [Owner ขอ 2026-08-10]
HISTORY_DAYS = int(os.environ.get("ML2R_HISTORY_DAYS", "7"))

BRANCH_CODE = "GRFML2"   # รหัสสาขาสำหรับเลขที่เอกสาร (GRF=ใบเบิก + ML2=สาขา)
BRANCH_NAME = "ML2"

# หมวดที่เป็น "บริการ" ไม่ใช่สินค้า -> ตัดออกจากใบเบิก (เพิ่มชื่อหมวดได้ตามต้องการ)
EXCLUDE_CATEGORIES = {"โรงแรมแมว"}

# หน้าต่างสำหรับคำนวณคะแนน ABC (ยอดขายย้อนหลังกี่วัน) - ใช้เพื่อจัดลำดับความสำคัญเท่านั้น
ABC_WINDOW_DAYS = 90


def next_doc_no(day):
    """สร้างเลขที่เอกสาร GRFML2-YYYYMMDD-NNN โดยรันเลขต่อรายวัน (เก็บใน .doc_seq.json)."""
    ymd = day.replace("-", "")
    seq_map = {}
    if os.path.exists(SEQ_FILE):
        try:
            with open(SEQ_FILE, "r", encoding="utf-8") as f:
                seq_map = json.load(f)
        except (ValueError, OSError):
            seq_map = {}
    seq = int(seq_map.get(ymd, 0)) + 1
    seq_map[ymd] = seq
    with open(SEQ_FILE, "w", encoding="utf-8") as f:
        json.dump(seq_map, f, ensure_ascii=False, indent=2)
    return f"{BRANCH_CODE}-{ymd}-{seq:03d}"


def find_latest_backup(folder):
    """หาไฟล์ Backup-*.db ที่ 'ใหม่ที่สุด' จากเวลาในชื่อไฟล์ (เผื่อ mtime ไม่ตรง)."""
    files = glob.glob(os.path.join(folder, "Backup-*.db"))
    if not files:
        raise FileNotFoundError("ไม่พบไฟล์ Backup-*.db ในโฟลเดอร์นี้")

    def key(path):
        # ชื่อไฟล์รูปแบบ Backup-2026-7-27T20.37.8.db
        m = re.search(r"Backup-(\d+)-(\d+)-(\d+)T(\d+)\.(\d+)\.(\d+)", os.path.basename(path))
        if m:
            y, mo, d, hh, mm, ss = map(int, m.groups())
            try:
                return datetime(y, mo, d, hh, mm, ss)
            except ValueError:
                pass
        return datetime.fromtimestamp(os.path.getmtime(path))

    return max(files, key=key), key


def load_master_pack(path):
    """โหลดตาราง Master แปลงหน่วย. คืน dict: single_barcode -> {pack_barcode, qty_per_pack, pack_unit_name}."""
    mapping = {}
    if not os.path.exists(path):
        return mapping
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            sb = (row.get("single_barcode") or "").strip()
            if not sb or sb.startswith("#"):
                continue  # ข้ามแถวว่าง / แถวตัวอย่างที่ comment ไว้
            try:
                qpp = float(row.get("qty_per_pack") or 0)
            except ValueError:
                qpp = 0
            if qpp <= 0:
                continue  # ยังไม่กรอกจำนวนต่อแพ็ค -> ข้าม
            mapping[sb] = {
                "pack_barcode": (row.get("pack_barcode") or "").strip(),
                "qty_per_pack": qpp,
                "pack_unit_name": (row.get("pack_unit_name") or "").strip() or "แพ็ค",
            }
    return mapping


def norm_name(s):
    """ปรับชื่อสินค้าให้เทียบกันได้ โดยตัด 'ส่วนที่บอกขนาดแพ็ค' ท้ายชื่อออก

    เช่น "MEO ขนม (4)รสปู 1โหล" กับ "MEO ขนม (4)รสปู 1ลัง" ต้องถือเป็นสินค้าเดียวกัน
    (ต่างกันแค่ขนาดแพ็ค) ไม่ใช่คนละสินค้า."""
    s = str(s or "").lower()
    s = re.sub(r"\*\s*\d+.*$", "", s)                                  # ตัด "*12", "*6ถุง"
    s = re.sub(r"\s*\d*\s*(ลัง|โหล|แพ็ค|แพค|ถุง|กล่อง|ซอง|ชิ้น)\s*$", "", s)  # ตัด "1โหล", "1ลัง"
    return re.sub(r"[^0-9a-z฀-๿]", "", s)


def find_latest_ml3(folder):
    """หา backup ML3 ล่าสุด (ชื่อ 'ML3 YYYY-MM-DD.db') = คลังต้นทางที่จะเบิกไปเติม ML2."""
    files = glob.glob(os.path.join(folder, "ML3 *.db"))
    if not files:
        return None

    def key(p):
        m = re.search(r"ML3 (\d{4})-(\d{1,2})-(\d{1,2})", os.path.basename(p))
        if m:
            y, mo, d = map(int, m.groups())
            try:
                return datetime(y, mo, d)
            except ValueError:
                pass
        return datetime.fromtimestamp(os.path.getmtime(p))

    return max(files, key=key)


def load_ml3_stock(path):
    """โหลดสต๊อกคงเหลือ ML3 -> dict: barcode -> Qty (จำนวนชิ้นในคลัง ML3)."""
    stock = {}
    if not path or not os.path.exists(path):
        return stock
    con = sqlite3.connect(path)
    for bc, q in con.execute("SELECT Barcode, Qty FROM Product WHERE IsDelete=0"):
        if bc is not None:
            stock[str(bc).strip()] = q or 0
    con.close()
    return stock


def apply_ml3_availability(requisition, ml3_stock):
    """เช็คใบเบิก ML2 กับสต๊อก ML3: ปรับจำนวนเท่าที่ ML3 มี + แยกตัวที่ ML3 หมดออก.

    คืน (รายการเบิกที่ ML3 จ่ายได้, รายการที่ต้องสั่งซื้อเข้า ML3)."""
    keep, purchase = [], []
    for r in requisition:
        bc = r["single_barcode"]
        # จำนวนที่อยากเบิกเป็นชิ้นจริง = total_pieces (คำนวณถูกต้องทุกกรณีแล้วใน build_requisition)
        # [แก้บั๊ก 2026-08-10 — Owner จับได้: เดิมคูณ req_qty*pack_mult ซ้ำ ทำให้แถวที่
        #  "เบิกเป็นชิ้นแต่ pack_mult ยังติดมา" (เช่น x20 ขายช้า) ถูกคิดเกิน 20 เท่า
        #  → ZOI CAT ขาย 7 ถูกคิดเป็น 140 → เห็น ML3 มี 113 → สั่งเบิก 113 ถุงผิดๆ]
        need_pcs = int(r.get("total_pieces") or r["req_qty"])
        avail = ml3_stock.get(bc)

        if avail is None or avail <= 0:              # ML3 ไม่มี/หมด -> ต้องสั่งซื้อเข้า
            r["ml3_stock"] = "ไม่มีสินค้า" if avail is None else "0"
            purchase.append(r)
            continue

        if avail >= need_pcs:                        # ML3 มีพอ -> เบิกตามเดิม
            r["ml3_status"] = "พอเบิก"
        else:                                        # ML3 มีไม่พอ -> เบิกเท่าที่มี (เป็นชิ้น)
            r["ml3_status"] = f"มีบางส่วน"
            r["req_qty"] = int(avail)
            r["req_unit"] = r.get("base_unit") or r["req_unit"]
            r["pack_mult"] = ""
            r["total_pieces"] = int(avail)           # เบิกเป็นชิ้นแล้ว รวมชิ้น = จำนวนเดียวกัน
            # เบิกเป็นชิ้นแล้ว ต้องโชว์บาร์โค้ดชิ้น ไม่ใช่บาร์โค้ดแพ็ค [ทีมคลังเคาะ 2026-08-10]
            r["req_barcode"] = r["single_barcode"]
            r["convert_note"] = (r.get("convert_note", "") + f" | ML3 มี {avail:g} เบิกเท่าที่มี").strip(" |")
        r["ml3_stock"] = f"{avail:g}"
        keep.append(r)
    return keep, purchase


def load_master_multiplier(path):
    """โหลด Master_Multiplier.xlsx -> dict: single_barcode -> [แถวแพ็คทั้งหมดของสินค้านั้น].

    ไฟล์นี้เก็บ 1 แถว = 1 แพ็ค ดังนั้นสินค้าตัวเดียวมีได้หลายแถว (โหล/ลัง)."""
    bysingle = {}
    if not os.path.exists(path):
        return bysingle
    try:
        import openpyxl
    except ImportError:
        print("  !! ไม่มีโมดูล openpyxl จึงข้ามไฟล์ Master_Multiplier.xlsx (ติดตั้ง: pip install openpyxl)")
        return bysingle

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Master"] if "Master" in wb.sheetnames else wb.worksheets[0]
    for idx, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx == 1:
            continue  # หัวตาราง
        # คอลัมน์: [0]NO. [1]category [2]pack_barcode [3]product_name [4]single_barcode [5]multiplier
        single = str(r[4]).strip() if r[4] not in (None, "") else ""
        if not single:
            continue
        try:
            mult = int(float(str(r[5]).strip()))
        except (TypeError, ValueError, IndexError):
            continue
        if mult <= 1:
            continue
        bysingle.setdefault(single, []).append({
            "row": idx,
            "pack_barcode": str(r[2]).strip() if r[2] else "",
            "name": str(r[3] or "").strip(),
            "mult": mult,
            "category": str(r[1] or "").strip(),
        })
    wb.close()
    return bysingle


def resolve_pack(barcode, pos_name, bysingle):
    """เลือกแพ็คที่จะใช้เบิกสำหรับบาร์โค้ดนี้ -> (แถวที่เลือก | None, สถานะ).

    กติกา (ยึดบาร์โค้ดเป็นตัวจริง, เอาชื่อ+ตัวคูณจากมาสเตอร์):
      - ไม่เจอในมาสเตอร์          -> ไม่มีแพ็ค เบิกเป็นชิ้น
      - เจอแถวเดียว                -> ใช้แถวนั้นเลย
      - หลายแถวแต่เป็นสินค้าเดียวกัน -> เลือกแพ็คเล็กสุด (โหลมาก่อนลัง)
      - หลายแถวและชื่อคนละสินค้า    -> ถ้าชื่อ POS ชี้ชัดได้ใช้แถวนั้น ไม่งั้นกันไว้ให้คนตรวจ
    """
    cands = bysingle.get(barcode)
    if not cands:
        return None, "ไม่มีในมาสเตอร์"
    if len(cands) == 1:
        return cands[0], "มาสเตอร์แถวเดียว"

    names = {norm_name(c["name"]) for c in cands}
    if len(names) == 1:
        return min(cands, key=lambda c: c["mult"]), "หลายขนาด-เลือกเล็กสุด"

    exact = [c for c in cands if norm_name(c["name"]) == norm_name(pos_name)]
    if len(exact) == 1:
        return exact[0], "ชื่อ POS ชี้ชัด"
    if len(exact) > 1:
        return min(exact, key=lambda c: c["mult"]), "หลายขนาด-เลือกเล็กสุด"
    return None, "ข้อมูลมาสเตอร์ขัดกัน"


def ensure_master_template(path):
    """สร้างไฟล์ CSV Master เปล่า (พร้อมหัวคอลัมน์ + ตัวอย่าง comment) ถ้ายังไม่มี."""
    if os.path.exists(path):
        return False
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["single_barcode", "product_name", "pack_barcode", "qty_per_pack", "pack_unit_name"])
        # แถวตัวอย่าง (ขึ้นต้น single_barcode ด้วย # = โปรแกรมจะข้าม ให้ลบ # เมื่อใช้จริง)
        w.writerow(["#8859601044263", "PRAMY AD1 ทูน่าในเจลลี่ 70g (ตัวอย่าง)", "8859601044260", "12", "โหล"])
        w.writerow(["#8850477009202", "MEO-G ลูกแมว 1kg (ตัวอย่าง-ไม่มีแพ็ค ให้ปล่อยว่าง)", "", "", ""])
    return True


def get_latest_sale_date(con):
    row = con.execute(
        "SELECT MAX(date(Complete)) FROM Orders WHERE IsDelete=0 AND date(Complete) > '2000-01-01'"
    ).fetchone()
    return row[0]


def get_target_day(con, backup_date):
    """เลือก 'วันที่สมบูรณ์ล่าสุด' สำหรับออกรายงาน.

    บทเรียน: ต้องใช้วันที่ที่ backup ถ่าย 'หลังร้านปิด' แล้วเท่านั้น ยอดถึงจะครบ.
    - ถ้าวันขายล่าสุด = วันเดียวกับ backup -> วันนั้นยังขายไม่จบ -> ถอยไปวันก่อนหน้า
    - ถ้าวันขายล่าสุด < วันที่ backup -> วันนั้นจบแล้ว ใช้ได้เลย
    คืน (target_day, is_complete)."""
    latest = get_latest_sale_date(con)
    if latest is None:
        return None, False
    if latest < backup_date:            # backup ถ่ายวันถัดมา -> วันขายล่าสุดจบแล้ว
        return latest, True
    # วันขายล่าสุด = วันเดียวกับ backup -> ยังไม่จบ ถอยไปวันก่อนที่มีการขาย
    prev = con.execute(
        "SELECT MAX(date(Complete)) FROM Orders WHERE IsDelete=0 AND date(Complete) < ?",
        (backup_date,)
    ).fetchone()[0]
    return prev, (prev is not None)


def query_today_sales(con, day):
    """ยอดขายรายสินค้าของวัน `day` (รวมทุกบิลที่ไม่ถูกลบ). คืน list ของ dict."""
    q = """
    SELECT od.Barcode                         AS barcode,
           COALESCE(p.Name, '(ไม่พบชื่อสินค้า)') AS name,
           COALESCE(cat.Name, '-')            AS category,
           COALESCE(u.Name, '-')              AS unit,
           SUM(od.Qty - COALESCE(od.QtyReturn,0)) AS qty_sold,
           SUM((od.Qty - COALESCE(od.QtyReturn,0)) * od.Price) AS amount,
           p.Qty                              AS stock_now,
           p.Left                             AS reorder_point
    FROM OrdersDetail od
    JOIN Orders o        ON o.Id = od.OrderId AND o.IsDelete = 0
    LEFT JOIN Product p  ON p.Barcode = od.Barcode AND p.IsDelete = 0
    LEFT JOIN ProductCategory cat ON cat.Id = p.Category
    LEFT JOIN ProductUnit u       ON u.Id = p.Unit
    WHERE date(o.Complete) = ?
      AND od.IsDelete = 0
    GROUP BY od.Barcode
    HAVING qty_sold > 0
    ORDER BY qty_sold DESC
    """
    cols = ["barcode", "name", "category", "unit", "qty_sold", "amount", "stock_now", "reorder_point"]
    return [dict(zip(cols, r)) for r in con.execute(q, (day,))]


def compute_abc(con, day, window_days):
    """คำนวณคลาส ABC ต่อบาร์โค้ด จากมูลค่าขายย้อนหลัง window_days วัน (Pareto 80/15/5)."""
    q = """
    SELECT od.Barcode AS barcode,
           SUM((od.Qty - COALESCE(od.QtyReturn,0)) * od.Price) AS amount
    FROM OrdersDetail od
    JOIN Orders o ON o.Id = od.OrderId AND o.IsDelete = 0
    WHERE od.IsDelete = 0
      AND date(o.Complete) > date(?, ?)
      AND date(o.Complete) <= ?
    GROUP BY od.Barcode
    HAVING amount > 0
    ORDER BY amount DESC
    """
    rows = con.execute(q, (day, f"-{window_days} days", day)).fetchall()
    total = sum(a for _, a in rows) or 1.0
    abc, cum = {}, 0.0
    for barcode, amount in rows:
        cum += amount
        share = cum / total
        abc[barcode] = "A" if share <= 0.80 else ("B" if share <= 0.95 else "C")
    return abc


def compute_avg_daily(con, day, window_days=COVER_WINDOW_DAYS):
    """ยอดขายเฉลี่ยต่อวัน (ชิ้น) ต่อบาร์โค้ด ย้อนหลัง window_days วัน — ใช้คำนวณ 'ของพอขายอีกกี่วัน'."""
    q = """
    SELECT od.Barcode, SUM(od.Qty - COALESCE(od.QtyReturn,0)) * 1.0 / ?
    FROM OrdersDetail od
    JOIN Orders o ON o.Id = od.OrderId AND o.IsDelete = 0
    WHERE od.IsDelete = 0
      AND date(o.Complete) > date(?, ?)
      AND date(o.Complete) <= ?
    GROUP BY od.Barcode
    """
    return {bc: avg for bc, avg in
            con.execute(q, (window_days, day, f"-{window_days} days", day))}


# เกณฑ์ "เคยขายดี-ของหมด" (ตาม Owner: กันสินค้าขายดีหายเพราะ ML3 ไม่มีของ)
RECOVERY_WINDOW_DAYS = 30   # ดูยอดขายย้อนหลังกี่วัน
RECOVERY_MIN_QTY = 10       # ต้องเคยขายอย่างน้อยกี่ชิ้นในช่วงนั้น


def build_recovery_list(con, day, bysingle, window=RECOVERY_WINDOW_DAYS, threshold=RECOVERY_MIN_QTY):
    """หาสินค้า 'เคยขายดีแต่ตอนนี้ของหมด และวันนี้ไม่ได้ขาย' -> ควรตามของจาก ML3.

    จุดบอดของระบบที่คิดจากยอดขายอย่างเดียว: ของหมด -> ขาย 0 -> ไม่เข้าใบเบิก -> หายถาวร."""
    q = """
    WITH hist AS (
      SELECT od.Barcode bc, SUM(od.Qty-COALESCE(od.QtyReturn,0)) q
      FROM OrdersDetail od JOIN Orders o ON o.Id=od.OrderId AND o.IsDelete=0
      WHERE date(o.Complete) > date(?, ?) AND date(o.Complete) <= ? AND od.IsDelete=0
      GROUP BY od.Barcode),
    h90 AS (
      SELECT od.Barcode bc, SUM(od.Qty-COALESCE(od.QtyReturn,0)) q
      FROM OrdersDetail od JOIN Orders o ON o.Id=od.OrderId AND o.IsDelete=0
      WHERE date(o.Complete) > date(?, '-90 days') AND date(o.Complete) <= ? AND od.IsDelete=0
      GROUP BY od.Barcode),
    today AS (
      SELECT DISTINCT od.Barcode bc FROM OrdersDetail od
      JOIN Orders o ON o.Id=od.OrderId AND o.IsDelete=0
      WHERE date(o.Complete)=? AND od.IsDelete=0)
    SELECT p.Barcode, p.Name, COALESCE(c.Name,'-') cat,
           hist.q q30, COALESCE(h90.q,0) q90, p.Qty, date(p.LastSaleTime) last_sale
    FROM hist
    JOIN Product p ON p.Barcode=hist.bc AND p.IsDelete=0
    LEFT JOIN h90 ON h90.bc=hist.bc
    LEFT JOIN ProductCategory c ON c.Id=p.Category
    WHERE p.Qty <= 0
      AND hist.q >= ?
      AND hist.bc NOT IN (SELECT bc FROM today)
      AND (c.Name IS NULL OR c.Name NOT IN ({}))
    ORDER BY hist.q DESC
    """.format(",".join("'%s'" % x for x in EXCLUDE_CATEGORIES))
    out = []
    for r in con.execute(q, (day, f"-{window} days", day, day, day, day, threshold)):
        bc, name, cat, q30, q90, stock, last_sale = r
        pack, _ = resolve_pack(bc, name, bysingle)
        disp_name = pack["name"] if (pack and pack.get("name")) else name
        if pack:
            # แนะนำเบิกกลับ 1 แพ็ค เป็นจุดเริ่ม (procurement ปรับได้)
            sugg_qty, sugg_unit = 1, PACK_UNIT_BY_MULT.get(pack["mult"], "แพ็ค")
        else:
            sugg_qty, sugg_unit = max(1, math.ceil(q30 / 4)), "ชิ้น"  # ~1 สัปดาห์
        out.append({
            "barcode": bc, "name": disp_name, "category": cat,
            "sold_30d": f"{q30:g}", "sold_90d": f"{q90:g}",
            "stock_now": f"{stock:g}", "last_sale": last_sale or "-",
            "pack_mult": pack["mult"] if pack else "",
            "suggest_qty": sugg_qty, "suggest_unit": sugg_unit,
        })
    return out


def build_requisition(today_sales, pack_map, bysingle, abc, avg_daily, avg_recent):
    """สร้างรายการเบิกจากยอดขายวันนี้.

    pack_map = ไฟล์ CSV แก้มือ (ใช้ก่อนเสมอ), bysingle = Master_Multiplier.xlsx
    avg_daily/avg_recent = ยอดขายเฉลี่ย/วัน 30 วัน / 7 วันล่าสุด (กฎ "ของยังพอ ไม่ต้องเบิก")
    คืน (รายการเบิก, รายการที่ข้อมูลมาสเตอร์ขัดกัน, รายการที่ข้ามเพราะของยังพอ)."""
    out, issues, skipped = [], [], []
    for s in today_sales:
        if s["category"] in EXCLUDE_CATEGORIES:   # ตัดบริการ (โรงแรมแมว ฯลฯ) ออกจากใบเบิก
            continue
        bc = s["barcode"]
        sold = s["qty_sold"] or 0
        need_units = max(1, math.ceil(sold))  # ขายกี่ชิ้น เบิกเท่านั้น ขั้นต่ำ 1

        # ด่านแรก — กฎ "ของยังพอ ไม่ต้องเบิก" [Owner เคาะ 3 วัน + มติทีมคลัง 2026-08-10]:
        # คงเหลือ ML2 พอขายเกินเกณฑ์ (A/B=3 วัน, C=2 วัน) -> ไม่เบิก ปล่อยขายกองเดิมออกก่อน
        # ฤดูกาล: ใช้ค่าเฉลี่ยที่แรงกว่าระหว่าง 30 วัน กับ 7 วันล่าสุด (ขาขึ้นเบิกไวขึ้นเอง)
        # ตาข่ายนิรภัย: ถ้าของเหลือไม่พอรองรับยอดขายซ้ำแบบวันล่าสุด (ขายพุ่งวันเดียว
        # แต่ค่าเฉลี่ยยังต่ำ) -> ต้องเบิกเสมอ ห้ามข้าม
        stock_now = s["stock_now"] if s["stock_now"] is not None else 0
        avg = max(avg_daily.get(bc, 0) or 0, avg_recent.get(bc, 0) or 0)
        max_cover = REQ_COVER_DAYS_BY_ABC.get(abc.get(bc, "-"), REQ_MAX_COVER_DAYS)
        if (stock_now > sold and
                stock_now > 0 and avg > 0 and (stock_now / avg) > max_cover):
            skipped.append({
                "single_barcode": bc, "name": s["name"], "category": s["category"],
                "abc": abc.get(bc, "-"),
                "qty_sold_today": f"{sold:g}", "stock_now": f"{stock_now:g}",
                "avg_daily": f"{avg:.2f}",
                "cover_days": f"{stock_now / avg:.0f}",
                "max_cover_days": max_cover,
            })
            continue

        # 1) ไฟล์ CSV แก้มือมาก่อน  2) แล้วค่อยดู Master_Multiplier.xlsx
        override = pack_map.get(bc)
        if override:
            pack = {"pack_barcode": override["pack_barcode"], "name": "",
                    "mult": int(override["qty_per_pack"])}
            unit_name = override["pack_unit_name"]
            status = "แก้มือ (CSV)"
        else:
            pack, status = resolve_pack(bc, s["name"], bysingle)
            unit_name = PACK_UNIT_BY_MULT.get(pack["mult"], "แพ็ค") if pack else ""

        if status == "ข้อมูลมาสเตอร์ขัดกัน":
            issues.append({
                "single_barcode": bc,
                "pos_name": s["name"],
                "master_rows": " | ".join(
                    f"แถว{c['row']} x{c['mult']} {c['name']}" for c in bysingle.get(bc, [])),
            })

        # ชื่อที่จะใช้ในเอกสาร: ยึดชื่อจากมาสเตอร์ ถ้าไม่มีค่อยใช้ชื่อ POS
        disp_name = pack["name"] if (pack and pack.get("name")) else s["name"]

        if pack:
            if pack["mult"] == 12:
                # กฎยกโหล x12: ขายดี(>=7) เบิกหลายโหล / ช้าเบิกเมื่อของ<4 / ช้า+ของพอ = ไม่เบิก
                stock_val = s["stock_now"] if s["stock_now"] is not None else 0
                if sold >= DOZEN_FAST_SALE:
                    pack_qty = max(1, math.ceil(need_units / 12))
                    note = f"ขายดี {sold:g} -> {pack_qty} โหล"
                elif stock_val < DOZEN_LOW_STOCK:
                    pack_qty = 1
                    note = f"ช้า แต่สต๊อกเหลือ {stock_val:g}(<4) -> 1 โหล"
                else:
                    continue   # ขายช้า + ของยังพอ -> ไม่เบิก (กันของกองเกิน)
                req_qty = pack_qty
                req_unit = unit_name
                req_barcode = pack["pack_barcode"] or bc
                total_pieces = pack_qty * 12
            elif pack["mult"] == 20:
                # กฎยกแพ็ค x20 [Piyawan ยืนยัน 2026-08-04]: ขายดี(>=8) ปัดเป็นแพ็คเต็ม /
                # ช้าแต่ของใกล้หมด(<=3) เบิก 1 แพ็คกันขาด / ช้า+ของพอ เบิกเป็นชิ้นเท่าที่ขายจริง (ไม่ยกแพ็ค)
                stock_val = s["stock_now"] if s["stock_now"] is not None else 0
                if sold >= X20_FAST_SALE:
                    pack_qty = max(1, math.ceil(need_units / 20))
                    req_qty = pack_qty
                    req_unit = unit_name
                    req_barcode = pack["pack_barcode"] or bc
                    total_pieces = pack_qty * 20
                    note = f"ขายดี {sold:g} -> {pack_qty} {unit_name}"
                elif stock_val <= X20_LOW_STOCK:
                    req_qty = 1
                    req_unit = unit_name
                    req_barcode = pack["pack_barcode"] or bc
                    total_pieces = 20
                    note = f"ช้า แต่สต๊อกเหลือ {stock_val:g}(<={X20_LOW_STOCK}) -> 1 {unit_name}"
                else:
                    # ขายช้า + ของยังพอ -> เบิกเป็นชิ้น ไม่ปัดเป็นแพ็ค (แพ็คใหญ่ กันของกองเกิน)
                    req_qty = need_units
                    req_unit = s["unit"]
                    req_barcode = bc
                    total_pieces = need_units
                    note = f"ช้า {sold:g} ชิ้น -> เบิกเป็นชิ้น (ไม่ยกแพ็ค x20 กันของกองเกิน)"
            else:
                pack_qty = math.ceil(need_units / pack["mult"])
                note = f"{need_units} {s['unit']} -> {pack_qty} {unit_name} (1 {unit_name} = {pack['mult']})"
                req_qty = pack_qty
                req_unit = unit_name
                req_barcode = pack["pack_barcode"] or bc
                total_pieces = pack_qty * pack["mult"]
        else:
            req_qty = need_units
            req_unit = s["unit"]
            req_barcode = bc
            total_pieces = need_units
            note = ""

        stock = s["stock_now"]
        reorder = s["reorder_point"] or 0
        # ธง 2 ระดับ:
        #   ด่วน   = สต็อกหมด/ติดลบ (<=0)
        #   ควรเติม = ต่ำกว่าหรือเท่าจุดสั่งซื้อขั้นต่ำ (Left) ที่ตั้งไว้ใน POS
        low_flag = ""
        if stock is not None and stock <= 0:
            low_flag = "ด่วน-ของหมด"
        elif stock is not None and reorder > 0 and stock <= reorder:
            low_flag = "ควรเติม"

        out.append({
            "req_barcode": req_barcode,
            "single_barcode": bc,
            "name": disp_name,
            "pos_name": s["name"],
            "category": s["category"],
            "abc": abc.get(bc, "-"),
            "qty_sold_today": f"{sold:g}",
            "stock_now": "" if stock is None else f"{stock:g}",
            "req_qty": req_qty,
            "req_unit": req_unit,
            "total_pieces": total_pieces,   # รวมเป็นชิ้นจริง — กันพนักงานสับสนตัวคูณ
            "base_unit": s["unit"],
            "pack_mult": pack["mult"] if pack else "",
            "match_status": status,
            "priority_flag": low_flag,
            "convert_note": note,
        })

    # เรียง: หมวด (เรียงชื่อ) -> ด่วนก่อน -> ยอดขายมากก่อน
    # [มติทีมคลัง 2026-08-10: ตัดการจัดกลุ่มคลาส ABC ออกจากใบ — เดินหยิบตามหมวดง่ายกว่า
    #  ABC ยังคำนวณอยู่เบื้องหลัง ใช้ในกฎ "ของยังพอ ไม่ต้องเบิก" + เก็บใน CSV]
    flag_rank = {"ด่วน-ของหมด": 0, "ควรเติม": 1, "": 2}
    out.sort(key=lambda r: (r["category"],
                            flag_rank.get(r["priority_flag"], 2),
                            -float(r["qty_sold_today"])))
    return out, issues, skipped


def write_csv(path, rows, headers):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        for r in rows:
            w.writerow({h: r.get(h, "") for h in headers})


def _day_th(day):
    try:
        return datetime.strptime(day, "%Y-%m-%d").strftime("%d/%m/%Y")
    except (ValueError, TypeError):
        return str(day)


def build_requisition_html(path, rows, doc_no, day, printed_at,
                           day_options=None, historical=False):
    """สร้างเอกสารใบเบิกสินค้าแบบพิมพ์ได้ (HTML -> Ctrl+P บันทึกเป็น PDF ได้).
    ซ่อนคอลัมน์ priority_flag / convert_note ตามที่ตกลง (ยังเก็บใน CSV).

    day_options = รายการวันที่ (ใหม่->เก่า) สำหรับดรอปดาวน์เลือกวันย้อนหลังบนเว็บ
    (ตัวแรกสุด = ล่าสุด ลิงก์ไป /requisition, วันอื่นไป /requisition-YYYY-MM-DD)
    historical = ใบย้อนหลัง (คำนวณใหม่จากสต๊อกปัจจุบัน อาจต่างจากใบที่พิมพ์วันนั้นจริง)"""
    day_th = _day_th(day)

    # ดรอปดาวน์เลือกวันที่ — โชว์เฉพาะตอนเปิดผ่านเว็บ (ซ่อนตอนพิมพ์ + ตอนเปิดเป็นไฟล์ local)
    daysel = ""
    if day_options:
        opts = []
        for i, d in enumerate(day_options):
            href = "/requisition" if i == 0 else f"/requisition-{d}"
            sel = " selected" if d == day else ""
            opts.append(f'<option value="{href}"{sel}>{_day_th(d)}</option>')
        daysel = ('<div class="daysel">📅 เลือกวันที่: '
                  f'<select onchange="location.href=this.value">{"".join(opts)}</select></div>')
    hist_note = ('<div class="histnote">⚠ ใบย้อนหลัง — คำนวณใหม่จากสต๊อกปัจจุบัน '
                 'ตัวเลขอาจต่างจากใบที่พิมพ์เช้าวันนั้นเล็กน้อย</div>') if historical else ""

    # [มติทีมคลัง 2026-08-10] ไม่คั่นหัวข้อคลาส ABC แล้ว — คั่นเฉพาะหมวดสินค้า
    body_rows = []
    i = 0
    cur_cat = None
    for r in rows:
        # หัวข้อคั่นเมื่อขึ้นหมวดใหม่
        if r["category"] != cur_cat:
            cur_cat = r["category"]
            body_rows.append(
                f"<tr class='hcat'><td colspan='10'>{html.escape(cur_cat or '-')}</td></tr>"
            )
        i += 1
        pack_txt = f"x{r['pack_mult']}" if r.get("pack_mult") else "-"
        ml3_txt = html.escape(str(r.get("ml3_stock", "-")))
        # "รวมเป็นชิ้น" = จำนวนชิ้นจริงที่ต้องหยิบ — พนักงานไม่ต้องคูณตัวคูณเอง
        pieces = r.get("total_pieces", r["req_qty"])
        body_rows.append(
            "<tr>"
            f"<td class='c'>{i}</td>"
            f"<td>{html.escape(str(r['req_barcode']))}</td>"
            f"<td>{html.escape(r['name'])}</td>"
            f"<td class='c'>{html.escape(str(r['qty_sold_today']))}</td>"
            f"<td class='c'>{html.escape(str(r['stock_now']))}</td>"
            f"<td class='c b'>{html.escape(str(r['req_qty']))}</td>"
            f"<td class='c'>{html.escape(r['req_unit'])}</td>"
            f"<td class='c'>{pack_txt}</td>"
            f"<td class='c b pieces'>{html.escape(str(pieces))}</td>"
            f"<td class='c'>{ml3_txt}</td>"
            "</tr>"
        )
    rows_html = "\n".join(body_rows)

    doc = f"""<!doctype html>
<html lang="th"><head><meta charset="utf-8">
<title>{html.escape(doc_no)}</title>
<style>
  /* บังคับกระดาษ A4 + ระยะขอบ; เนื้อหาเกินจะไหลไปหน้า 2-3-4 อัตโนมัติ */
  @page {{ size: A4 portrait; margin: 12mm 10mm 14mm 10mm; }}
  * {{ box-sizing: border-box; }}
  body {{ font-family: "Tahoma","TH Sarabun New",sans-serif; color:#111; margin:0; font-size:13px; }}
  h1 {{ text-align:center; font-size:20px; margin:0 0 4px; }}
  .meta {{ display:flex; justify-content:space-between; margin:8px 0 4px; }}
  .meta div {{ line-height:1.7; }}
  .docno {{ font-weight:bold; }}
  table {{ width:100%; border-collapse:collapse; margin-top:10px; }}
  th, td {{ border:1px solid #888; padding:3px 6px; }}
  thead th {{ background:#f0f0f0; font-size:12px; }}
  td.c {{ text-align:center; }}
  td.b {{ font-weight:bold; }}
  /* คอลัมน์ "รวมเป็นชิ้น" เน้นพื้นเหลือง — เลขที่พนักงานต้องหยิบจริง ไม่ต้องคูณเอง */
  td.pieces {{ background:#fdf0c2; -webkit-print-color-adjust:exact; print-color-adjust:exact; }}
  tr.hcat td {{ background:#eef3f7; font-weight:bold; padding-left:16px; }}
  .total {{ margin-top:8px; font-weight:bold; }}
  .sign {{ display:flex; justify-content:space-around; margin-top:40px; text-align:center; }}
  .sign div {{ width:30%; }}
  .line {{ border-top:1px dotted #333; margin-top:36px; padding-top:4px; }}
  .daysel {{ text-align:right; font-size:13px; margin:0 0 6px; }}
  .daysel select {{ font-family:inherit; font-size:13px; padding:3px 6px; }}
  .histnote {{ background:#fff3e0; border:1px solid #e0a960; border-radius:4px;
               padding:4px 8px; font-size:12px; margin:4px 0; }}

  /* มุมมองบนจอ: ทำให้ดูเหมือนแผ่น A4 */
  @media screen {{
    body {{ background:#e9e9e9; }}
    .sheet {{ width:210mm; min-height:297mm; margin:10px auto; padding:12mm 10mm 14mm;
             background:#fff; box-shadow:0 1px 6px rgba(0,0,0,.25); }}
  }}

  /* ตอนพิมพ์: หัวตารางซ้ำทุกหน้า + ไม่ตัดแถว/หัวข้อ/ช่องเซ็นขาดครึ่ง */
  @media print {{
    .sheet {{ width:auto; margin:0; padding:0; box-shadow:none; }}
    thead {{ display:table-header-group; }}
    tr {{ page-break-inside:avoid; break-inside:avoid; }}
    tr.hcat {{ page-break-after:avoid; break-after:avoid; }}
    .sign {{ page-break-inside:avoid; break-inside:avoid; }}
    .daysel {{ display:none; }}
  }}
</style></head>
<body>
  <div class="sheet">
  {daysel}
  <h1>เอกสารใบเบิกสินค้า</h1>
  {hist_note}
  <div class="meta">
    <div>สาขา: <b>{html.escape(BRANCH_NAME)}</b><br>วันที่ขาย: <b>{day_th}</b></div>
    <div style="text-align:right">เลขที่: <span class="docno">{html.escape(doc_no)}</span><br>พิมพ์เมื่อ: {printed_at}</div>
  </div>
  <table>
    <thead><tr>
      <th>ลำดับ</th><th>บาร์โค้ด</th><th>สินค้า</th>
      <th>ขายวันนี้</th><th>คงเหลือ</th><th>จำนวนเบิก</th><th>หน่วย</th><th>ต่อแพ็ค</th><th>รวมเป็นชิ้น</th><th>ML3 มี</th>
    </tr></thead>
    <tbody>
{rows_html}
    </tbody>
  </table>
  <p class="total">รวมทั้งสิ้น {len(rows)} รายการ</p>
  <div class="sign">
    <div><div class="line">ผู้ขอเบิก</div></div>
    <div><div class="line">ผู้จ่ายสินค้า</div></div>
    <div><div class="line">ผู้รับสินค้า</div></div>
  </div>
  </div>
  <script>/* เปิดเป็นไฟล์ local (file://) ลิงก์ /requisition ใช้ไม่ได้ -> ซ่อนดรอปดาวน์ */
  if(location.protocol.indexOf("http")!==0){{var d=document.querySelector(".daysel");if(d)d.style.display="none";}}</script>
</body></html>"""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write(doc)


def build_day_requisition(con, day, pack_map, bysingle, ml3_stock, doc_no):
    """คำนวณรายการเบิกของวัน `day` ด้วยแกนเดียวกับใบหลัก — ใช้สร้างใบย้อนหลัง.

    หมายเหตุ: สต๊อก (ML2/ML3) ในไฟล์ backup เป็นค่าปัจจุบัน ใบย้อนหลังจึงเป็นการ
    คำนวณใหม่จากสต๊อกวันนี้ อาจต่างจากใบที่พิมพ์เช้าวันนั้นจริงเล็กน้อย."""
    abc = compute_abc(con, day, ABC_WINDOW_DAYS)
    avg_daily = compute_avg_daily(con, day)
    avg_recent = compute_avg_daily(con, day, COVER_RECENT_DAYS)
    sales = query_today_sales(con, day)
    rows, _issues, _skipped = build_requisition(sales, pack_map, bysingle, abc, avg_daily, avg_recent)
    if ml3_stock:
        rows, _purchase = apply_ml3_availability(rows, ml3_stock)
    for r in rows:
        r["doc_no"] = doc_no
    return rows


def main():
    db_path, keyfn = find_latest_backup(BACKUP_DIR)
    backup_dt = keyfn(db_path)                    # เวลาที่ backup ถ่าย (จากชื่อไฟล์)
    backup_date = backup_dt.strftime("%Y-%m-%d")
    print("=" * 64)
    print("ML2 Daily Report  |  ไฟล์ backup ล่าสุด:")
    print("  ", os.path.basename(db_path), f"(ถ่าย {backup_dt.strftime('%d/%m %H:%M')})")

    created = ensure_master_template(MASTER_CSV)
    if created:
        print("  สร้างไฟล์ Master เปล่าให้แล้ว: master_pack_mapping.csv")

    con = sqlite3.connect(db_path)

    day, complete = get_target_day(con, backup_date)
    doc_no = next_doc_no(day)
    printed_at = datetime.now().strftime("%d/%m/%Y %H:%M")
    print(f"  วันที่ประมวลผล (วันสมบูรณ์ล่าสุด): {day}")
    if not complete:
        print("  !! เตือน: ยังไม่พบ backup ที่ถ่ายหลังวันนี้ปิดร้าน ยอดอาจยังไม่ครบ")
    print("  เลขที่เอกสารใบเบิก:", doc_no)
    print("=" * 64)

    pack_map = load_master_pack(MASTER_CSV)
    bysingle = load_master_multiplier(MASTER_XLSX)
    print(f"  Master_Multiplier.xlsx: สินค้า {len(bysingle)} บาร์โค้ด "
          f"({sum(len(v) for v in bysingle.values())} แถวแพ็ค)")

    today_sales = query_today_sales(con, day)
    abc = compute_abc(con, day, ABC_WINDOW_DAYS)
    avg_daily = compute_avg_daily(con, day)                       # เฉลี่ย 30 วัน
    avg_recent = compute_avg_daily(con, day, COVER_RECENT_DAYS)   # เฉลี่ย 7 วันล่าสุด (จับฤดูกาล)
    requisition, issues, skipped = build_requisition(
        today_sales, pack_map, bysingle, abc, avg_daily, avg_recent)

    # เช็คสต๊อก ML3 (คลังต้นทาง): ปรับจำนวนเท่าที่มี + แยกตัวที่ต้องสั่งซื้อเข้า
    ml3_path = find_latest_ml3(ML3_DIR)
    purchase = []
    ml3_stock = None
    if ml3_path:
        ml3_stock = load_ml3_stock(ml3_path)
        print(f"  สต๊อก ML3: {os.path.basename(ml3_path)} ({len(ml3_stock):,} รายการ)")
        requisition, purchase = apply_ml3_availability(requisition, ml3_stock)
    else:
        print("  !! ไม่พบ backup ML3 -> ข้ามการเช็คสต๊อกต้นทาง")

    for r in requisition + purchase:   # แนบเลขที่เอกสารทุกแถว (ใช้ตอนส่งขึ้น Supabase ภายหลัง)
        r["doc_no"] = doc_no
    recovery = build_recovery_list(con, day, bysingle)   # เคยขายดี-ของหมด (ตาม Owner)

    # ---- รายงาน 1: ขายวันนี้ ----
    total_qty = sum(s["qty_sold"] for s in today_sales)
    total_amt = sum(s["amount"] or 0 for s in today_sales)
    print(f"\n[1] วันนี้ขายออกไป {len(today_sales)} รายการ | รวม {total_qty:g} หน่วย | ยอด {total_amt:,.0f} บาท")
    print("-" * 64)
    print(f"{'จำนวน':>7}  {'สินค้า'}")
    for s in today_sales[:15]:
        print(f"{s['qty_sold']:>7g}  {s['name'][:48]}")
    if len(today_sales) > 15:
        print(f"     ... และอีก {len(today_sales) - 15} รายการ (ดูเต็มใน output/today_sales.csv)")

    # ---- รายงาน 2: ต้องเบิก ----
    print(f"\n[2] ใบเบิกสินค้า {doc_no} | {len(requisition)} รายการ "
          f"(ขายกี่ชิ้น เบิกเท่านั้น ขั้นต่ำ 1 | ของพอขายเกิน {REQ_MAX_COVER_DAYS} วันไม่เบิก)")
    print(f"    ข้ามเพราะของยังพอ (เกิน {REQ_MAX_COVER_DAYS} วันขาย): {len(skipped)} รายการ "
          f"-> output/skipped_enough_stock.csv")
    from collections import Counter
    print("    ผลการจับคู่มาสเตอร์:")
    for st, c in Counter(r["match_status"] for r in requisition).most_common():
        print(f"       {st:24} {c:>4} รายการ")
    renamed = sum(1 for r in requisition if r["name"] != r["pos_name"])
    print(f"    ใช้ชื่อจากมาสเตอร์แทนชื่อ POS: {renamed} รายการ")
    print("-" * 64)
    print(f"{'ABC':<4}{'เบิก':>6} {'หน่วย':<8}{'ต่อแพ็ค':<8}{'สินค้า'}")
    for r in requisition[:20]:
        pk = f"x{r['pack_mult']}" if r["pack_mult"] else "-"
        print(f"{r['abc']:<4}{r['req_qty']:>6} {r['req_unit']:<8}{pk:<8}{r['name'][:36]}")
    if len(requisition) > 20:
        print(f"     ... และอีก {len(requisition) - 20} รายการ (ดูเต็มในเอกสาร)")

    # ---- เขียนไฟล์ ----
    write_csv(os.path.join(OUTPUT_DIR, "today_sales.csv"), today_sales,
              ["barcode", "name", "category", "unit", "qty_sold", "amount", "stock_now", "reorder_point"])
    # CSV ข้อมูล: เก็บ priority_flag / convert_note ไว้ (ไม่โชว์ในเอกสาร แต่เก็บไว้ใช้ต่อ)
    write_csv(os.path.join(OUTPUT_DIR, "requisition.csv"), requisition,
              ["doc_no", "req_barcode", "single_barcode", "name", "pos_name", "category", "abc",
               "qty_sold_today", "stock_now", "req_qty", "req_unit", "total_pieces", "pack_mult",
               "ml3_stock", "match_status", "priority_flag", "convert_note"])
    # รายการที่ข้ามเพราะของยังพอ (กฎ 3 วัน, C=2 วัน) -> ไว้ตรวจย้อน/ปรับเกณฑ์
    write_csv(os.path.join(OUTPUT_DIR, "skipped_enough_stock.csv"), skipped,
              ["single_barcode", "name", "category", "abc", "qty_sold_today",
               "stock_now", "avg_daily", "cover_days", "max_cover_days"])
    # ต้องสั่งซื้อเข้า ML3 (ML2 ขายออก แต่ ML3 ก็หมด) -> ให้ฝ่ายจัดซื้อตามซัพพลายเออร์
    write_csv(os.path.join(OUTPUT_DIR, "purchase_ml3.csv"), purchase,
              ["single_barcode", "name", "category", "abc", "qty_sold_today",
               "stock_now", "ml3_stock", "pack_mult"])
    # รายชื่อวันย้อนหลังสำหรับดรอปดาวน์เลือกวันที่ (ใหม่ -> เก่า, ตัวแรก = วันล่าสุด)
    hist_days = [r[0] for r in con.execute(
        "SELECT DISTINCT date(Complete) d FROM Orders WHERE IsDelete=0 "
        "AND date(Complete) <= ? AND date(Complete) > '2000-01-01' "
        "ORDER BY d DESC LIMIT ?", (day, HISTORY_DAYS))]

    # เอกสารใบเบิกแบบพิมพ์ได้ (ซ่อน priority_flag / convert_note)
    doc_path = os.path.join(OUTPUT_DIR, "requisition_document.html")
    build_requisition_html(doc_path, requisition, doc_no, day, printed_at,
                           day_options=hist_days)

    # ---- ใบเบิกย้อนหลัง (เลือกวันที่บนเว็บ) [Owner ขอ 2026-08-10] ----
    for d in hist_days:
        if d == day:
            rows_d, doc_d, hist = requisition, doc_no, False
        else:
            doc_d = f"{BRANCH_CODE}-{d.replace('-', '')}-001"
            rows_d = build_day_requisition(con, d, pack_map, bysingle, ml3_stock, doc_d)
            hist = True
        build_requisition_html(
            os.path.join(OUTPUT_DIR, "req_days", f"requisition-{d}.html"),
            rows_d, doc_d, d, printed_at, day_options=hist_days, historical=hist)
    print(f"  ใบเบิกย้อนหลัง {len(hist_days)} วัน -> output/req_days/ (ดรอปดาวน์เลือกวันที่บนเว็บ)")
    # รายการที่ข้อมูลมาสเตอร์ขัดกัน -> ไว้ให้ไปแก้ต้นทาง
    write_csv(os.path.join(OUTPUT_DIR, "master_issues.csv"), issues,
              ["single_barcode", "pos_name", "master_rows"])
    # เคยขายดี-ของหมด (ตาม Owner) -> ทบทวนสัปดาห์ละครั้ง ตามของจาก ML3
    write_csv(os.path.join(OUTPUT_DIR, "recovery_watchlist.csv"), recovery,
              ["barcode", "name", "category", "sold_30d", "sold_90d", "stock_now",
               "last_sale", "pack_mult", "suggest_qty", "suggest_unit"])

    # ---- รายงาน 2.5: ต้องสั่งซื้อเข้า ML3 ----
    if purchase:
        print(f"\n[2.5] ต้องสั่งซื้อเข้า ML3 {len(purchase)} รายการ (ML2 ขายออก แต่ ML3 ก็หมด)")
        print("-" * 64)
        print(f"{'ขายวันนี้':>8}  {'สินค้า'}")
        for r in sorted(purchase, key=lambda x: -float(x["qty_sold_today"]))[:15]:
            print(f"{str(r['qty_sold_today']):>8}  {r['name'][:44]}")
        if len(purchase) > 15:
            print(f"     ... และอีก {len(purchase) - 15} รายการ (ดูใน output/purchase_ml3.csv)")

    # ---- รายงาน 3: เคยขายดี-ของหมด ----
    print(f"\n[3] เคยขายดี-ของหมด {len(recovery)} รายการ "
          f"(ขาย >= {RECOVERY_MIN_QTY} ใน {RECOVERY_WINDOW_DAYS} วัน แต่ตอนนี้ของหมด | ตามของจาก ML3)")
    print("-" * 64)
    print(f"{'30วัน':>6}{'90วัน':>6}  {'ขายล่าสุด':<12}{'สินค้า'}")
    for r in recovery:
        print(f"{r['sold_30d']:>6}{r['sold_90d']:>6}  {str(r['last_sale']):<12}{r['name'][:38]}")

    print("\n" + "=" * 64)
    print("บันทึกไฟล์แล้ว:")
    print("  output/requisition_document.html  <-- ใบเบิกสินค้า (เปิดเบราว์เซอร์ Ctrl+P พิมพ์ได้)")
    print("  output/requisition.csv            (ข้อมูลดิบ เก็บครบทุกคอลัมน์)")
    print("  output/today_sales.csv            (ขายวันนี้ทั้งหมด)")
    print(f"  output/purchase_ml3.csv           ({len(purchase)} รายการต้องสั่งซื้อเข้า ML3)")
    print(f"  output/skipped_enough_stock.csv   ({len(skipped)} รายการข้าม-ของยังพอเกิน {REQ_MAX_COVER_DAYS} วัน)")
    print(f"  output/master_issues.csv          ({len(issues)} รายการที่มาสเตอร์ขัดกัน ต้องไปแก้)")
    print(f"  output/recovery_watchlist.csv     ({len(recovery)} รายการเคยขายดี-ของหมด | ดูสัปดาห์ละครั้ง)")
    print("=" * 64)


if __name__ == "__main__":
    main()
