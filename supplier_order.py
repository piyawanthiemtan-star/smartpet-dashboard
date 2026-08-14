# -*- coding: utf-8 -*-
"""หน้าสั่งซื้อรายซัพพลายเออร์ (/purchasing-order) — เครื่องมือของทีมจัดซื้อ (เบียร์)
[Owner สั่งทำ 2026-08-13]

เลือกซัพ → เห็นสินค้าของเจ้านั้น (สต๊อก/ขายเฉลี่ย/แนะนำสั่ง/ทุนซื้อ) → ติ๊ก+แก้จำนวน
→ พิมพ์ใบสั่งซื้อ (พร้อมเงื่อนไขชำระ+บัญชีโอนจาก POS) หรือคัดลอกข้อความส่ง LINE ให้เซลล์

แหล่งข้อมูล (backup ML3 — โกดังเป็นคนสั่งซื้อเข้า):
- Vendor: ซัพ 43 เจ้า (ชื่อ/เบอร์/เครดิต/หมายเหตุที่มีเลขบัญชีโอน)
- สินค้า→ซัพ: อนุมานจากประวัติรับของ (ImportProduct ล่าสุดที่ระบุซัพ) + ใบ PO เดิม
  (Product.VendorId ใน POS ว่าง 0% — ใช้ไม่ได้; แผนที่นี้โตเองเมื่อทีมเลือกซัพตอนรับของ)
- แนะนำสั่ง = เติมให้พอขาย 14 วัน: ceil(ขายเฉลี่ย30วัน × 14 − สต๊อก) ขั้นต่ำ 0
- นโยบายข้อมูล: มี "ทุนซื้อ" ได้ (งานจัดซื้อ) แต่ไม่มีราคาขาย/มาร์จิ้น

รัน: python supplier_order.py
env override (GitHub Actions): SUP_ML3_DIR, SUP_TPL, SUP_OUT
"""
import datetime
import glob
import json
import math
import os
import re
import sqlite3
import sys
from collections import Counter, defaultdict

sys.stdout.reconfigure(encoding="utf-8")
HERE = os.path.dirname(os.path.abspath(__file__))
BASE = r"D:\1. SmartPet AI Framework"

ML3_DIR = os.environ.get("SUP_ML3_DIR", BASE + r"\SmartPetData\Import\OnePoint\Blackup 2026")
ML2_DIR = os.environ.get("SUP_ML2_DIR", BASE + r"\SmartPetBackup\Daily\Onepoint ML2 Blackup")
TPL = os.environ.get("SUP_TPL", os.path.join(HERE, "supplier_order_template.html"))
OUT = os.environ.get("SUP_OUT", os.path.join(HERE, "output", "purchasing-order.html"))
MASTER = os.environ.get("SUP_MASTER", os.path.join(HERE, "Master_Multiplier.xlsx"))
VMAP_MANUAL = os.environ.get("SUP_VMAP", os.path.join(HERE, "vendor_map_manual.csv"))
VENDOR_INFO = os.environ.get("SUP_VINFO", os.path.join(HERE, "vendor_info.csv"))

BKK = datetime.timezone(datetime.timedelta(hours=7))   # ตรึง +07:00 — GitHub runner เป็น UTC (บทเรียน 29 ก.ค.)
COVER_DAYS = 14   # แนะนำสั่ง = เติมให้พอกี่วัน (เบียร์แก้จำนวนเองได้อยู่แล้ว)
PACK_UNIT_BY_MULT = {12: "โหล", 24: "ลัง"}   # ชื่อหน่วยแพ็คตามตัวคูณ (อื่นๆ = "ลัง")

# หัวเอกสารใบสั่งซื้อมาตรฐาน [Owner สั่ง 2026-08-13] — ที่อยู่/เลขผู้เสียภาษี Owner ส่งมาเติมได้
COMPANY = {
    "name": "บริษัท เลิฟเพ็ท โกลบอลพลัส จำกัด",
    "branch": "โกดังเลยสมาร์ทเพ็ท (ML3) อ.เมือง จ.เลย",
    "buyer": "นฤมล วงศ์นอก (เบียร์) ฝ่ายจัดซื้อ",
    "approver": "ปิยวรรณ เทียมทัน (Owner)",
}


def load_master_mult(path):
    """Master_Multiplier.xlsx -> บาร์โค้ดเดี่ยว -> (ตัวคูณเล็กสุด, ชื่อหน่วยแพ็ค)
    ใช้บอกจัดซื้อว่าตัวไหนต้อง 'สั่งเต็มลังเท่านั้น' (1 ลัง/โหล = กี่ชิ้น)."""
    out = {}
    if not os.path.exists(path):
        return out
    try:
        import openpyxl
    except ImportError:
        return out
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Master"] if "Master" in wb.sheetnames else wb.worksheets[0]
    for idx, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx == 1:
            continue
        single = str(r[4]).strip() if r[4] not in (None, "") else ""
        try:
            mult = int(float(str(r[5]).strip()))
        except (TypeError, ValueError, IndexError):
            continue
        if not single or mult <= 1:
            continue
        if single not in out or mult < out[single][0]:
            out[single] = (mult, PACK_UNIT_BY_MULT.get(mult, "ลัง"))
    wb.close()
    return out


def load_manual_map(path):
    """vendor_map_manual.csv -> บาร์โค้ด -> vendor id
    การผูกมือจากหน้าเว็บ (เบียร์ติ๊กแล้วบันทึก) — ชนะทุกชั้น รวมถึงใช้แก้ตัวที่ระบบเดาผิด"""
    out = {}
    if not os.path.exists(path):
        return out
    import csv
    with open(path, encoding="utf-8-sig", newline="") as f:
        for row in csv.reader(f):
            if len(row) >= 2 and row[0].strip() and row[0].strip().lower() != "barcode":
                out[row[0].strip()] = row[1].strip()
    return out


def norm_vendor_name(s):
    """ชื่อบริษัทแบบเทียบได้: ตัดคำนำหน้า/ต่อท้ายนิติบุคคล + วรรค + จุด (ไว้จับคู่ตอนไม่มีเลขผู้เสียภาษี)"""
    s = str(s or "").strip().lower()
    for w in ("ห้างหุ้นส่วนจำกัด", "หจก.", "บริษัท", "บจก.", "บ.", "(สำนักงานใหญ่)", "(มหาชน)", "จำกัด"):
        s = s.replace(w, "")
    return re.sub(r"[\s\.]+", "", s)


def load_vendor_info(path):
    """vendor_info.csv — ข้อมูลซัพจากบัญชี (Owner ส่ง 14 ส.ค. 2569):
    เลขผู้เสียภาษี (=Vendor.Id ใน POS) · หมวดสินค้า · โทร/แฟกซ์ · ธนาคาร+เลขบัญชีโอน · เงื่อนไขชำระ"""
    rows = []
    if not os.path.exists(path):
        return rows
    import csv
    with open(path, encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            r = {k: (v or "").strip() for k, v in r.items()}
            if r.get("name"):
                rows.append(r)
    return rows


def newest_db(folder):
    fs = glob.glob(os.path.join(folder, "*.db"))
    return max(fs, key=os.path.getmtime) if fs else None


def main():
    db = newest_db(ML3_DIR)
    if not db:
        sys.exit(f"ไม่พบไฟล์ backup ใน {ML3_DIR}")
    con = sqlite3.connect(f"file:{db}?mode=ro", uri=True)
    print("ใช้ backup:", os.path.basename(db))

    # ความสดของข้อมูล = บิลใบล่าสุดใน backup (กฎเหล็ก: ห้ามใช้เวลารันสคริปต์/mtime)
    last_bill = con.execute("SELECT MAX([Create]) FROM Orders WHERE IsDelete=0").fetchone()[0]
    last_bill_th = ""
    if last_bill:
        try:
            last_bill_th = datetime.datetime.fromisoformat(
                str(last_bill).replace("T", " ")).strftime("%d/%m/%Y %H:%M")
        except ValueError:
            pass

    # --- ซัพพลายเออร์ ---
    vendors = {}
    for vid, name, tel, remarks, credit in con.execute(
            "SELECT Id, Name, Tel, Remarks, CreditDay FROM Vendor WHERE IsDelete=0"):
        vendors[str(vid).strip()] = {
            "id": str(vid).strip(), "name": (name or "").strip(),
            "tel": (tel or "").strip(), "note": (remarks or "").strip(),
            "credit": str(credit or "").strip(), "vcat": "",
        }
    print(f"ซัพพลายเออร์: {len(vendors)} เจ้า")

    # --- เติมข้อมูลจากบัญชี (vendor_info.csv): บัญชีโอน/เครดิตเทอม/โทร/หมวด ---
    # จับคู่ด้วยเลขผู้เสียภาษี (= Vendor.Id ใน POS) ก่อน · ไม่มีเลขค่อยเทียบชื่อแบบตัดคำนิติบุคคล
    info_rows = load_vendor_info(VENDOR_INFO)
    by_tax = {r["tax_id"]: r for r in info_rows if r.get("tax_id")}
    by_name = {norm_vendor_name(r["name"]): r for r in info_rows}
    matched_info = set()
    for v in vendors.values():
        r = by_tax.get(v["id"]) or by_name.get(norm_vendor_name(v["name"]))
        if not r:
            continue
        matched_info.add(id(r))
        if not v["tel"] and r.get("tel"):
            v["tel"] = r["tel"]
        v["vcat"] = r.get("category", "")
        pay = []
        m = re.search(r"เครดิต\s*(\d+)", r.get("terms", ""))
        if m:
            if not v["credit"] or v["credit"] == "0":
                v["credit"] = m.group(1)
        elif r.get("terms"):
            pay.append("ชำระ" + r["terms"])          # เงินสด / คิวอาร์โค้ด
        if r.get("bank") and r.get("account"):
            acct_digits = re.sub(r"\D", "", r["account"])
            if acct_digits and acct_digits not in re.sub(r"\D", "", v["note"]):
                pay.append("โอน " + r["bank"] + " " + r["account"])
        if pay:
            v["note"] = (v["note"] + " · " if v["note"] else "") + " · ".join(pay)
    unmatched = [r["name"] for r in info_rows if id(r) not in matched_info]
    print(f"ข้อมูลบัญชีซัพ: จับคู่ได้ {len(matched_info)}/{len(info_rows)}"
          + (f" · จับคู่ไม่ได้: {', '.join(unmatched)}" if unmatched else ""))

    # --- แผนที่ สินค้า -> ซัพ (ประวัติรับของล่าสุดชนะ · PO เดิมเป็นตัวเสริม) ---
    vmap = {}   # bc -> vendor id
    for bc, vid in con.execute("""
        SELECT d.Barcode, i.VenderId FROM ImportProductDetail d
        JOIN ImportProduct i ON i.Id = d.ImportId AND i.IsDelete=0
        WHERE d.IsDelete=0 AND TRIM(COALESCE(i.VenderId,'')) <> ''
        ORDER BY i.[Create] ASC"""):
        vmap[str(bc).strip()] = str(vid).strip()          # แถวหลัง (ใหม่กว่า) ทับแถวแรก
    for bc, vid in con.execute("""
        SELECT d.Barcode, o.VendorId FROM PurchaseOrderDetail d
        JOIN PurchaseOrder o ON o.Id = d.PurchaseOrderId AND o.IsDelete=0
        WHERE d.IsDelete=0 AND TRIM(COALESCE(o.VendorId,'')) <> ''"""):
        vmap.setdefault(str(bc).strip(), str(vid).strip())  # ไม่ทับของจากประวัติรับของ
    print(f"สินค้าที่โยงซัพได้: {len(vmap):,} บาร์โค้ด")

    # --- ชั้นผูกมือจากหน้าเว็บ (ชนะทุกชั้น) — ตัดตัวที่ชี้ซัพที่ถูกลบไปแล้วทิ้ง ---
    manual = {bc: v for bc, v in load_manual_map(VMAP_MANUAL).items() if v in vendors}
    vmap.update(manual)
    print(f"ผูกมือจากหน้าเว็บ: {len(manual):,} บาร์โค้ด")

    # --- ยอดขาย 30/90 วัน: รวม 2 สาขา (ML3 + ML2) [Owner สั่ง 2026-08-13] ---
    # ดีมานด์จริงของการสั่งซื้อเข้าโกดัง = ขายส่งที่ ML3 + ขายปลีกที่ ML2 (ไม่นับโอนระหว่างสาขา จึงไม่ซ้ำ)
    sold30, sold90 = {}, {}
    stock_ml2 = {}   # สต๊อกคงเหลือฝั่งหน้าร้าน [Owner เคาะ 2026-08-13: โชว์ ML3/ML2/รวม]
    sale_dbs = [("ML3", con)]
    ml2_db = newest_db(ML2_DIR)
    if ml2_db:
        sale_dbs.append(("ML2", sqlite3.connect(f"file:{ml2_db}?mode=ro", uri=True)))
        print("รวมยอดขาย+สต๊อก ML2 จาก:", os.path.basename(ml2_db))
    else:
        print("!! ไม่พบ backup ML2 — ยอดขาย/สต๊อกเป็นของ ML3 สาขาเดียว")
    for _br, c in sale_dbs:
        for days, box in ((30, sold30), (90, sold90)):
            for bc, q in c.execute(f"""
                SELECT d.Barcode, SUM(d.Qty-COALESCE(d.QtyReturn,0))
                FROM OrdersDetail d JOIN Orders o ON o.Id=d.OrderId AND o.IsDelete=0
                WHERE d.IsDelete=0 AND date(o.[Create]) > date('now','-{days} day')
                GROUP BY d.Barcode"""):
                k = str(bc).strip()
                box[k] = box.get(k, 0) + (q or 0)
    if ml2_db:
        for bc, q in sale_dbs[1][1].execute(
                "SELECT Barcode, Qty FROM Product WHERE IsDelete=0"):
            if bc is not None:
                stock_ml2[str(bc).strip()] = q or 0
        sale_dbs[1][1].close()

    # --- ตัวคูณจาก Master: ตัวไหนต้องสั่งเต็มลัง/โหลเท่านั้น ---
    mults = load_master_mult(MASTER)
    print(f"ตัวคูณจาก Master: {len(mults):,} บาร์โค้ด")

    # --- อ่านสินค้าทั้งหมดก่อน (ใช้สร้างชั้นเดาซัพจากประเภท/แบรนด์) ---
    # p.VendorId = ซัพจากการ์ดสินค้าใน POS (ทีมเริ่มกรอก 14 ส.ค. 2569 — เดิมว่าง 0%)
    raw = [(str(r[0]).strip(), r[1] or "", r[2] or 0, r[3] or 0, r[4] or "", r[5] or "-",
            str(r[6] or "").strip())
           for r in con.execute("""
        SELECT p.Barcode, p.Name, p.Qty, p.Cost, COALESCE(u.Name,''), COALESCE(c.Name,'-'),
               p.VendorId
        FROM Product p LEFT JOIN ProductUnit u ON u.Id = p.Unit
        LEFT JOIN ProductCategory c ON c.Id = p.Category
        WHERE p.IsDelete=0""")]

    # ลำดับชั้นข้อมูลซัพ: ผูกมือบนเว็บ > การ์ดสินค้า POS > ประวัติรับของ/PO (> เดาประเภท/แบรนด์)
    n_pos = 0
    def direct_vid(bc, pos_vid):
        if bc in manual:
            return manual[bc]
        if pos_vid and pos_vid in vendors:
            return pos_vid
        return vmap.get(bc, "")

    # --- ชั้นเดาซัพ [Owner เคาะ 2026-08-13: "จับจากรหัสประเภทได้เลย ไม่ต้องรอ"] ---
    # ของ 2,206 ตัวไม่เคยถูกบันทึกซัพ (ทีมรับผ่านเครื่องไม่เลือกบริษัท) จึงเดาจาก:
    #  1) ประเภทสินค้า: สินค้าประเภทเดียวกันที่ผูกแล้ว >=80% ชี้เจ้าเดียว (ตย.>=2) -> ทั้งประเภทเจ้านั้น
    #  2) แบรนด์ (คำแรกของชื่อ): เกณฑ์เดียวกัน (ตย.>=3) — ตัวเสริมเมื่อประเภทไม่มีเบาะแส
    def brand_key(nm):
        s = re.sub(r"^[\(\)\[\]0-9A-Z\-\.]{0,8}\s*", "", str(nm or "").strip()).strip("()[] ")
        tok = re.split(r"[\s/]+", s)
        return (tok[0] if tok and tok[0] else str(nm or "").split(" ")[0]).lower()

    cat_v, brand_v = defaultdict(Counter), defaultdict(Counter)
    for bc, nm, _q, _c, _u, cat, pv in raw:
        v = direct_vid(bc, pv)
        if not v:
            continue
        if cat != "-":
            cat_v[cat][v] += 1
        brand_v[brand_key(nm)][v] += 1

    def dominant(counter, n_min):
        top, n = counter.most_common(1)[0]
        return top if (n >= n_min and n / sum(counter.values()) >= 0.8) else ""

    cat_guess = {c: dominant(cnt, 2) for c, cnt in cat_v.items()}
    brand_guess = {b: dominant(cnt, 3) for b, cnt in brand_v.items()}
    n_by_cat = n_by_brand = 0

    # --- สินค้า (เอาเฉพาะตัวที่มีความเคลื่อนไหว: มีสต๊อก หรือขายใน 90 วัน) ---
    # สต๊อกคงเหลือ = ของ ML3 (โกดัง — คลังที่สั่งซื้อเข้า) [Owner ยืนยัน 2026-08-13]
    items = []
    for bc, name, qty, cost, unit, cat, pv in raw:
        st3 = qty or 0
        st2 = stock_ml2.get(bc, 0)
        s30 = sold30.get(bc, 0)
        s90 = sold90.get(bc, 0)
        if st3 <= 0 and st2 <= 0 and s90 <= 0:
            continue
        avg = s30 / 30.0
        total = st3 + st2
        # แนะนำสั่ง = หักสต๊อกรวม 2 สาขา (สมมาตรกับดีมานด์รวม — กันสั่งเกินเพราะของกองที่ ML2)
        need_units = max(0, math.ceil(avg * COVER_DAYS - total))
        mult, punit = mults.get(bc, (0, ""))
        # มีตัวคูณ -> สั่งเต็มลังเท่านั้น: แนะนำเป็นจำนวนลังปัดขึ้น [Owner เคาะ 2026-08-13]
        sugg = math.ceil(need_units / mult) if (mult and need_units > 0) else need_units
        vid = direct_vid(bc, pv)
        src = "m" if bc in manual else ""   # แหล่งที่มา: m=ผูกมือ ""=การ์ดสินค้า/ประวัติรับของ/PO c=เดาประเภท b=เดาแบรนด์
        if vid and src != "m" and pv and pv in vendors:
            n_pos += 1
        if not vid:
            vid = cat_guess.get((cat or "-").strip(), "")
            if vid:
                n_by_cat += 1
                src = "c"
            else:
                vid = brand_guess.get(brand_key(name), "")
                if vid:
                    n_by_brand += 1
                    src = "b"
        items.append({
            "bc": bc, "name": (name or "").strip(), "unit": (unit or "").strip(),
            "cat": (cat or "-").strip(), "vid": vid, "src": src,
            "st3": round(st3, 1), "st2": round(st2, 1), "tot": round(total, 1),
            "s30": round(s30, 1),
            "cost": round(cost or 0, 2), "sugg": sugg,
            "mult": mult, "punit": punit,
        })
    n_mapped = sum(1 for x in items if x["vid"])
    print(f"สินค้าในหน้า: {len(items):,} (โยงซัพแล้ว {n_mapped:,} "
          f"[ผูกมือ {sum(1 for x in items if x['src'] == 'm'):,} "
          f"· การ์ดสินค้า POS {n_pos:,} "
          f"· เดาจากประเภท {n_by_cat:,} · จากแบรนด์ {n_by_brand:,}] "
          f"· ยังไม่ระบุ {len(items)-n_mapped:,} · มีตัวคูณ {sum(1 for x in items if x['mult']):,})")

    data = {
        "generatedTh": datetime.datetime.now(BKK).strftime("%d/%m/%Y %H:%M"),
        "lastBillTh": last_bill_th,
        "coverDays": COVER_DAYS,
        "company": COMPANY,
        "vendors": sorted(vendors.values(), key=lambda v: v["name"]),
        "items": items,
    }
    tpl = open(TPL, encoding="utf-8").read()
    html = tpl.replace("__DATA__", json.dumps(data, ensure_ascii=False))
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        f.write(html)
    print("เขียนไฟล์:", OUT, f"({len(html)/1024:.0f} KB)")


if __name__ == "__main__":
    main()
